from rest_framework import status, generics
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework_simplejwt.tokens import RefreshToken
from django.db.models import Count, Sum, Q, F, Max
from django.utils import timezone
from datetime import timedelta
import pandas as pd
from io import BytesIO
import traceback

from .models import PushNotificationDevice
from .serializers import (
    MobileLoginSerializer, PushNotificationDeviceSerializer,
    RegisterDeviceSerializer, ProductDashboardSerializer,
    ProductDashboardStatsSerializer, FrequentClientSerializer,
    ExcelImportSerializer, ExcelImportResultSerializer
)
from apps.products.models import Product, Category
from apps.clients.models import Client
from apps.sales.models import Sale, SaleItem


@api_view(['POST'])
@permission_classes([AllowAny])
def mobile_login(request):
    """Login específico para aplicación móvil"""
    import logging
    logger = logging.getLogger(__name__)
    
    # Log para debugging
    logger.info(f"Login attempt - Data received: {request.data}")
    
    serializer = MobileLoginSerializer(data=request.data)
    
    if not serializer.is_valid():
        logger.warning(f"Login validation failed: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        user = serializer.validated_data['user']
        
        # Generar tokens JWT
        refresh = RefreshToken.for_user(user)
        
        # Actualizar actividad del usuario
        user.update_activity(ip_address=get_client_ip(request))
        
        logger.info(f"Login successful for user: {user.username}")
        
        return Response({
            'refresh': str(refresh),
            'access': str(refresh.access_token),
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'full_name': user.full_name,
                'role': user.role
            }
        }, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Login error: {str(e)}")
        return Response({'error': 'Error interno del servidor'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def get_client_ip(request):
    """Obtener IP del cliente"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def register_push_device(request):
    """Registrar dispositivo para recibir notificaciones push"""
    import logging
    from django.db import transaction, OperationalError
    import time
    
    logger = logging.getLogger(__name__)
    
    serializer = RegisterDeviceSerializer(data=request.data)
    
    if serializer.is_valid():
        device_token = serializer.validated_data['device_token']
        device_type = serializer.validated_data['device_type']
        device_id = serializer.validated_data.get('device_id', '')
        app_version = serializer.validated_data.get('app_version', '')
        
        # Manejar "database is locked" con retry logic (común en SQLite con requests concurrentes)
        max_retries = 3
        retry_delay = 0.1  # 100ms
        
        for attempt in range(max_retries):
            try:
                # Usar transacción atómica para evitar locks
                with transaction.atomic():
                    device, created = PushNotificationDevice.objects.update_or_create(
                        user=request.user,
                        device_token=device_token,
                        defaults={
                            'device_type': device_type,
                            'device_id': device_id,
                            'app_version': app_version,
                            'is_active': True
                        }
                    )
                    
                    response_serializer = PushNotificationDeviceSerializer(device)
                    return Response(
                        response_serializer.data, 
                        status=status.HTTP_201_CREATED if created else status.HTTP_200_OK
                    )
                    
            except OperationalError as e:
                if 'database is locked' in str(e).lower() and attempt < max_retries - 1:
                    # Esperar un poco antes de reintentar
                    time.sleep(retry_delay * (attempt + 1))
                    logger.warning(f'Database locked, retry {attempt + 1}/{max_retries}')
                    continue
                else:
                    # Si es el último intento o no es un lock, re-lanzar el error
                    logger.error(f'Error registrando dispositivo: {e}')
                    return Response(
                        {'error': 'Error de base de datos. Intenta nuevamente.'},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR
                    )
            except Exception as e:
                logger.error(f'Error inesperado registrando dispositivo: {e}', exc_info=True)
                return Response(
                    {'error': str(e)},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
    
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def unregister_push_device(request, device_token):
    """Eliminar registro de dispositivo"""
    try:
        device = PushNotificationDevice.objects.get(
            user=request.user,
            device_token=device_token
        )
        device.is_active = False
        device.save()
        return Response({'message': 'Dispositivo desregistrado correctamente'}, status=status.HTTP_200_OK)
    except PushNotificationDevice.DoesNotExist:
        return Response({'error': 'Dispositivo no encontrado'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def products_dashboard(request):
    """Dashboard de productos para móvil"""
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        # Obtener productos con paginación - CRÍTICO: usar distinct() para evitar duplicados
        products = Product.objects.filter(is_active=True).select_related('category').distinct()
        
        # Filtros opcionales
        category_id = request.GET.get('category')
        if category_id:
            try:
                products = products.filter(category_id=int(category_id))
            except ValueError:
                pass
        
        search = request.GET.get('search')
        if search:
            products = products.filter(
                Q(name__icontains=search) |
                Q(sku__icontains=search) |
                Q(barcode__icontains=search)
            ).distinct()  # Asegurar no duplicados en búsqueda
        
        # Filtro de stock
        stock_filter = request.GET.get('stock_status')
        if stock_filter == 'low':
            products = products.filter(stock__lte=F('min_stock')).exclude(stock=0).distinct()
        elif stock_filter == 'out':
            products = products.filter(stock=0).distinct()
        
        # Ordenamiento - CRÍTICO: agregar 'id' al final para garantizar orden consistente
        ordering = request.GET.get('ordering', '-created_at')
        # Validar que el campo de ordenamiento sea válido
        valid_orderings = ['-created_at', 'created_at', '-name', 'name', '-price', 'price', '-stock', 'stock']
        if ordering not in valid_orderings:
            ordering = '-created_at'
        
        # Agregar 'id' al ordenamiento para garantizar orden consistente y evitar duplicados
        if 'id' not in ordering and '-id' not in ordering:
            if ordering.startswith('-'):
                products = products.order_by(ordering, '-id')
            else:
                products = products.order_by(ordering, 'id')
        else:
            products = products.order_by(ordering)
        
        # Contar total antes de paginar
        total_count = products.count()
        
        # Paginación
        try:
            page_size = int(request.GET.get('page_size', 20))
            page = int(request.GET.get('page', 1))
        except (ValueError, TypeError):
            page_size = 20
            page = 1
        
        if page_size < 1:
            page_size = 20
        if page < 1:
            page = 1
        
        start = (page - 1) * page_size
        end = start + page_size
        
        products_page = products[start:end]
        
        # Estadísticas - calcular de forma segura
        total_products = Product.objects.filter(is_active=True).count()
        
        # Calcular stock bajo y sin stock usando querysets
        out_of_stock = Product.objects.filter(is_active=True, stock=0).count()
        
        # Para stock bajo, usar values_list para evitar cargar todos los objetos
        low_stock_products = Product.objects.filter(is_active=True).exclude(stock=0).values_list('stock', 'min_stock')
        low_stock = sum(1 for stock, min_stock in low_stock_products if stock <= min_stock)
        
        # Calcular valor total del stock de forma eficiente
        products_with_cost = Product.objects.filter(
            is_active=True
        ).exclude(
            cost__isnull=True
        ).exclude(
            cost=0
        ).values_list('stock', 'cost')
        
        total_stock_value = sum(
            float(stock * cost) for stock, cost in products_with_cost if stock and cost
        )
        
        categories_count = Category.objects.filter(is_active=True).count()
        
        stats = {
            'total_products': total_products,
            'low_stock_products': low_stock,
            'out_of_stock_products': out_of_stock,
            'total_stock_value': round(total_stock_value, 2),
            'categories_count': categories_count
        }
        
        # Serializar productos con manejo de errores
        try:
            # CRÍTICO: Eliminar duplicados antes de serializar (por si acaso)
            seen_ids = set()
            unique_products = []
            for product in products_page:
                if product.id not in seen_ids:
                    seen_ids.add(product.id)
                    unique_products.append(product)
            
            logger.info(f'📊 Productos únicos a serializar: {len(unique_products)} de {len(products_page)} recibidos')
            
            serializer = ProductDashboardSerializer(unique_products, many=True)
            serializer_data = serializer.data
            
        except Exception as e:
            logger.error(f'Error serializando productos: {str(e)}')
            # Retornar productos básicos sin serializer
            seen_ids = set()
            unique_products = []
            for product in products_page:
                if product.id not in seen_ids:
                    seen_ids.add(product.id)
                    unique_products.append(product)
            
            serializer_data = [{
                'id': p.id,
                'name': p.name or 'Sin nombre',
                'sku': p.sku or 'N/A',
                'price': float(p.price) if p.price else 0,
                'stock': int(p.stock) if p.stock else 0,
                'min_stock': int(p.min_stock) if p.min_stock else 0,
                'max_stock': int(p.max_stock) if p.max_stock else 0,
                'category_name': p.category.name if p.category else 'Sin categoría',
                'stock_status': 'out_of_stock' if p.stock == 0 else ('low_stock' if p.stock <= p.min_stock else 'in_stock'),
                'is_low_stock': p.stock <= p.min_stock if p.min_stock else False,
                'is_active': p.is_active,
            } for p in unique_products]
        
        # CRÍTICO: Asegurar que no haya duplicados en la respuesta final
        final_results = []
        seen_result_ids = set()
        for item in serializer_data:
            item_id = item.get('id')
            if item_id and item_id not in seen_result_ids:
                seen_result_ids.add(item_id)
                # Asegurar que el stock sea coherente (entero, no negativo)
                if 'stock' in item:
                    item['stock'] = max(0, int(item['stock']) if item['stock'] else 0)
                final_results.append(item)
        
        logger.info(f'📦 Productos finales sin duplicados: {len(final_results)}')
        
        return Response({
            'results': final_results,
            'stats': stats,
            'pagination': {
                'page': page,
                'page_size': page_size,
                'total': total_count,
                'total_pages': (total_count + page_size - 1) // page_size if total_count > 0 else 0
            }
        })
    except Exception as e:
        logger.error(f'Error en products_dashboard: {str(e)}', exc_info=True)
        return Response({
            'error': 'Error al cargar productos',
            'detail': str(e) if request.user.is_staff else 'Contacta al administrador'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_excel_template(request):
    """Descargar template de Excel para importación de productos"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import io
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos Template"
    
    # Estilos
    header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    required_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
    optional_fill = PatternFill(start_color="f0f9ff", end_color="f0f9ff", fill_type="solid")
    
    # Encabezados - Columnas requeridas
    required_headers = ['name', 'sku', 'price', 'stock', 'category']
    optional_headers = ['description', 'cost', 'min_stock', 'max_stock', 'barcode', 'tags', 'is_digital']
    
    all_headers = required_headers + optional_headers
    
    # Escribir encabezados
    for col_num, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Marcar columnas requeridas
        if header in required_headers:
            cell.fill = required_fill
    
    # Ancho de columnas
    column_widths = {
        'name': 30, 'sku': 20, 'price': 15, 'stock': 12, 'category': 20,
        'description': 40, 'cost': 15, 'min_stock': 12, 'max_stock': 12,
        'barcode': 20, 'tags': 25, 'is_digital': 12
    }
    
    for col_num, header in enumerate(all_headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = column_widths.get(header, 15)
    
    # Agregar múltiples filas de ejemplo con diferentes productos
    example_rows = [
        # Fila 2: Producto físico
        [
            'Refrigerador Samsung 500L',  # name
            'SKU-REF-001',                 # sku
            15000.00,                      # price
            25,                            # stock
            'Electrodomésticos',           # category
            'Refrigerador de 500 litros, tecnología inverter',  # description
            12000.00,                      # cost
            5,                             # min_stock
            100,                           # max_stock
            '7501234567890',               # barcode
            'refrigerador, samsung, 500L', # tags
            False                          # is_digital
        ],
        # Fila 3: Producto físico con menos datos
        [
            'Lavadora LG 15kg',
            'SKU-LAV-002',
            8000.00,
            15,
            'Electrodomésticos',
            '',                            # description vacío (opcional)
            '',                            # cost vacío (opcional)
            '',                            # min_stock vacío (opcional)
            '',                            # max_stock vacío (opcional)
            '',                            # barcode vacío (opcional)
            '',                            # tags vacío (opcional)
            False
        ],
        # Fila 4: Producto digital
        [
            'Curso Online Python Avanzado',
            'SKU-CURSO-003',
            599.99,
            0,                             # stock 0 para digital
            'Cursos',
            'Curso completo de Python nivel avanzado con 50 horas de contenido',
            300.00,
            0,
            0,
            '',
            'curso, python, digital, online',
            True                           # is_digital = True
        ],
    ]
    
    # Escribir filas de ejemplo
    for row_index, example_row in enumerate(example_rows, start=2):
        for col_num, value in enumerate(example_row, 1):
            cell = ws.cell(row=row_index, column=col_num, value=value)
            # Aplicar estilo a las celdas de ejemplo
            cell.fill = optional_fill
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    excel_data = output.getvalue()
    output.close()
    
    # Crear respuesta HTTP con el archivo Excel
    response = HttpResponse(
        excel_data,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="productos_template.xlsx"'
    response['Content-Length'] = len(excel_data)
    
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def preview_excel(request):
    """Previsualizar archivo Excel antes de importar"""
    import logging
    logger = logging.getLogger(__name__)
    
    # Validar el archivo con el serializer
    serializer = ExcelImportSerializer(data=request.data)
    
    if not serializer.is_valid():
        logger.error(f'Errores de validación: {serializer.errors}')
        return Response({
            'error': 'Error de validación del archivo',
            'details': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    excel_file = serializer.validated_data['file']
    logger.info(f'Previsualizando archivo: {excel_file.name}, tamaño: {excel_file.size} bytes')
    
    try:
        # Leer archivo Excel
        if excel_file.name.endswith('.xlsx'):
            df = pd.read_excel(excel_file, engine='openpyxl')
        else:
            df = pd.read_excel(excel_file, engine='xlrd')
        
        # Obtener primeras 10 filas para previsualización
        preview_rows = df.head(10).fillna('').to_dict('records')
        
        # Validar columnas requeridas
        required_columns = ['name', 'sku', 'price', 'stock', 'category']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        # Información del archivo
        file_info = {
            'total_rows': len(df),
            'columns': list(df.columns),
            'required_columns': required_columns,
            'missing_columns': missing_columns,
            'has_required_columns': len(missing_columns) == 0,
            'preview_rows': preview_rows,
        }
        
        return Response(file_info, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f'Error al leer Excel: {str(e)}', exc_info=True)
        return Response({
            'error': f'Error al procesar archivo: {str(e)}',
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_products_excel(request):
    """Importar productos desde archivo Excel"""
    import logging
    import traceback
    from django.db import transaction
    logger = logging.getLogger(__name__)
    
    logger.info(f'🔔 ========== REQUEST RECIBIDO PARA IMPORTACIÓN ==========')
    logger.info(f'🔔 Usuario: {request.user.username} (ID: {request.user.id})')
    logger.info(f'🔔 Método: {request.method}')
    logger.info(f'🔔 Content-Type: {request.content_type}')
    logger.info(f'🔔 FILES keys: {list(request.FILES.keys())}')
    logger.info(f'🔔 DATA keys: {list(request.data.keys())}')
    
    # CRÍTICO: Asegurar que los logs se escriban inmediatamente
    import sys
    sys.stdout.flush()
    
    # En DRF, request.data automáticamente combina request.POST y request.FILES
    # Para serializers con archivos, simplemente pasamos request.data
    # DRF automáticamente incluye los archivos de request.FILES
    serializer = ExcelImportSerializer(data=request.data)
    
    if not serializer.is_valid():
        logger.error(f'❌ Errores de validación del serializer: {serializer.errors}')
        logger.error(f'❌ FILES disponibles: {list(request.FILES.keys())}')
        logger.error(f'❌ DATA keys: {list(request.data.keys())}')
        
        # Log detallado de lo que se recibió
        if request.FILES:
            for key, file in request.FILES.items():
                logger.error(f'❌ FILES[{key}]: name={file.name}, size={file.size}, content_type={getattr(file, "content_type", "N/A")}')
        
        return Response({
            'error': 'Error de validación del archivo',
            'details': serializer.errors,
            'received_files': list(request.FILES.keys()),
            'received_data': list(request.data.keys())
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Verificar que hay un archivo después de la validación
    if 'file' not in serializer.validated_data:
        logger.error('El archivo no está en validated_data después de la validación')
        return Response({
            'error': 'No se recibió el archivo. Asegúrate de enviar el archivo con la clave "file".',
            'received_files': list(request.FILES.keys()),
            'received_data': list(request.data.keys())
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Obtener el archivo validado del serializer
    excel_file = serializer.validated_data['file']
    logger.info(f'📥 ========== INICIANDO IMPORTACIÓN DE EXCEL ==========')
    logger.info(f'📥 Archivo recibido: {excel_file.name}')
    logger.info(f'📥 Tamaño: {excel_file.size} bytes')
    logger.info(f'📥 Content-Type: {getattr(excel_file, "content_type", "N/A")}')
    logger.info(f'📥 Usuario: {request.user.username} (ID: {request.user.id})')
    
    # Validar que el nombre del archivo termine en .xlsx o .xls
    file_name_lower = excel_file.name.lower() if excel_file.name else ''
    if not file_name_lower.endswith(('.xlsx', '.xls')):
        logger.error(f'❌ Nombre de archivo inválido: {excel_file.name}')
        return Response({
            'error': f'El archivo debe tener extensión .xlsx o .xls. Archivo recibido: {excel_file.name}',
        }, status=status.HTTP_400_BAD_REQUEST)
    
    errors = []
    successful = 0
    failed = 0
    
    try:
        logger.info(f'📖 Leyendo archivo Excel...')
        
        # Leer archivo Excel
        if excel_file.name.endswith('.xlsx'):
            df = pd.read_excel(excel_file, engine='openpyxl')
        else:
            df = pd.read_excel(excel_file, engine='xlrd')
        
        total_rows = len(df)
        logger.info(f'📊 Total de filas leídas: {total_rows}')
        logger.info(f'📊 Columnas encontradas: {list(df.columns)}')
        
        # Validar columnas requeridas
        required_columns = ['name', 'sku', 'price', 'stock', 'category']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            logger.error(f'❌ Columnas faltantes: {missing_columns}')
            return Response({
                'error': f'Columnas faltantes: {", ".join(missing_columns)}',
                'required_columns': required_columns
            }, status=status.HTTP_400_BAD_REQUEST)
        
        logger.info(f'✅ Todas las columnas requeridas están presentes')
        logger.info(f'🔄 Iniciando procesamiento de {total_rows} filas...')
        
        # Usar transacción atómica para asegurar que todos los cambios se guarden
        with transaction.atomic():
            # Procesar cada fila
            for index, row in df.iterrows():
                try:
                    # Validar datos requeridos primero
                    name = str(row.get('name', '')).strip()
                    sku = str(row.get('sku', '')).strip()
                    
                    # Saltar filas vacías o con instrucciones
                    if not name or not sku or pd.isna(row.get('name')) or pd.isna(row.get('sku')):
                        continue  # Ignorar filas vacías
                    
                    # Saltar filas que parecen ser instrucciones
                    if 'INSTRUCCIONES' in name.upper() or 'REQUERIDO' in name.upper() or 'OPCIONAL' in name.upper():
                        continue
                    
                    # Validar y convertir datos numéricos
                    try:
                        price_raw = row.get('price', 0)
                        stock_raw = row.get('stock', 0)
                        
                        # Convertir price
                        if pd.isna(price_raw) or price_raw == '':
                            price = 0
                        else:
                            price = float(price_raw)
                        
                        # Convertir stock - CRÍTICO: debe ser entero positivo
                        if pd.isna(stock_raw) or stock_raw == '':
                            stock = 0
                        else:
                            # Asegurar que sea entero
                            stock_float = float(stock_raw)
                            stock = max(0, int(stock_float))  # Asegurar que sea positivo
                        
                        logger.info(f"📝 Fila {index + 2}: SKU={sku}, Price={price}, Stock={stock} (raw: {stock_raw})")
                        
                    except (ValueError, TypeError) as e:
                        logger.error(f"❌ Error convirtiendo números en fila {index + 2}: {e}, price_raw={price_raw}, stock_raw={stock_raw}")
                        errors.append(f"Fila {index + 2}: price o stock no son números válidos (price={price_raw}, stock={stock_raw})")
                        failed += 1
                        continue
                    
                    if price <= 0:
                        errors.append(f"Fila {index + 2}: El precio debe ser mayor a 0")
                        failed += 1
                        continue
                    
                    # Obtener o crear categoría
                    category_name = str(row.get('category', 'Sin categoría')).strip()
                    if not category_name or pd.isna(row.get('category')):
                        category_name = 'Sin categoría'
                    
                    category, _ = Category.objects.get_or_create(
                        name=category_name,
                        defaults={'description': f'Categoría importada desde Excel'}
                    )
                    
                    # Procesar campos opcionales de forma segura
                    description = str(row.get('description', '')).strip() if pd.notna(row.get('description')) else ''
                    
                    try:
                        cost = float(row.get('cost')) if pd.notna(row.get('cost')) and str(row.get('cost')).strip() else price * 0.7
                    except (ValueError, TypeError):
                        cost = price * 0.7
                    
                    try:
                        min_stock = int(float(row.get('min_stock'))) if pd.notna(row.get('min_stock')) and str(row.get('min_stock')).strip() else 10
                    except (ValueError, TypeError):
                        min_stock = 10
                    
                    try:
                        max_stock = int(float(row.get('max_stock'))) if pd.notna(row.get('max_stock')) and str(row.get('max_stock')).strip() else 1000
                    except (ValueError, TypeError):
                        max_stock = 1000
                    
                    barcode = str(row.get('barcode', '')).strip() if pd.notna(row.get('barcode')) else ''
                    tags = str(row.get('tags', '')).strip() if pd.notna(row.get('tags')) else ''
                    
                    # Procesar is_digital de forma segura
                    is_digital_val = row.get('is_digital', False)
                    if pd.notna(is_digital_val):
                        if isinstance(is_digital_val, bool):
                            is_digital = is_digital_val
                        elif isinstance(is_digital_val, str):
                            is_digital = is_digital_val.lower() in ('true', '1', 'yes', 'verdadero')
                        else:
                            is_digital = bool(is_digital_val)
                    else:
                        is_digital = False
                    
                    # Verificar si el producto ya existe
                    try:
                        product = Product.objects.get(sku=sku)
                        created = False
                        
                        # ACTUALIZAR producto existente
                        stock_before = product.stock
                        logger.info(f"🔄 Actualizando producto existente: SKU={sku}, Stock actual={stock_before}, Stock a sumar={stock} (tipo: {type(stock)})")
                        
                        # Asegurar que stock sea int positivo
                        stock_to_add = int(max(0, stock))
                        
                        product.name = name
                        product.price = price
                        # SUMAR el stock nuevo al stock existente (no reemplazar)
                        product.stock = int(product.stock) + stock_to_add
                        product.category = category
                        
                        # Actualizar otros campos si se proporcionan
                        if description:
                            product.description = description
                        if cost and cost > 0:
                            product.cost = cost
                        if min_stock is not None:
                            product.min_stock = int(min_stock)
                        if max_stock is not None:
                            product.max_stock = int(max_stock)
                        if barcode:
                            product.barcode = barcode
                        if tags:
                            product.tags = tags
                        if pd.notna(row.get('is_digital')):
                            product.is_digital = is_digital
                        
                        # Guardar explícitamente todos los campos actualizados
                        # Usar update_fields para optimizar y asegurar que se guarden todos los cambios
                        fields_to_update = ['name', 'price', 'stock', 'category']
                        if description:
                            fields_to_update.append('description')
                        if cost and cost > 0:
                            fields_to_update.append('cost')
                        if min_stock is not None:
                            fields_to_update.append('min_stock')
                        if max_stock is not None:
                            fields_to_update.append('max_stock')
                        if barcode:
                            fields_to_update.append('barcode')
                        if tags:
                            fields_to_update.append('tags')
                        if pd.notna(row.get('is_digital')):
                            fields_to_update.append('is_digital')
                        
                        # Guardar con update_fields explícito
                        product.save(update_fields=fields_to_update)
                        
                        # CRÍTICO: Refrescar desde BD para verificar que se guardó
                        product.refresh_from_db()
                        
                        logger.info(f"✅ Producto actualizado: SKU={sku}, Stock: {stock_before} + {stock_to_add} = {product.stock}")
                        logger.info(f"   Verificado en BD después de refresh: stock={product.stock}, id={product.id}")
                        
                        # Verificación de integridad CRÍTICA
                        expected_stock = stock_before + stock_to_add
                        if product.stock != expected_stock:
                            logger.error(f"⚠️ PROBLEMA CRÍTICO: Stock no coincidente!")
                            logger.error(f"   Esperado: {expected_stock}, Obtenido: {product.stock}")
                            logger.error(f"   Stock antes: {stock_before}, Stock a sumar: {stock_to_add}")
                            
                            # Reintentar el guardado
                            product.stock = expected_stock
                            product.save(update_fields=['stock'])
                            product.refresh_from_db()
                            logger.info(f"   Reintentado guardado. Nuevo stock: {product.stock}")
                        else:
                            logger.info(f"✅ VERIFICACIÓN OK: Stock coincide correctamente ({product.stock})")
                        
                    except Product.DoesNotExist:
                        # CREAR nuevo producto
                        created = True
                        stock_initial = int(max(0, stock))
                        logger.info(f"➕ Creando nuevo producto: SKU={sku}, Stock inicial={stock_initial} (tipo: {type(stock_initial)})")
                        
                        product = Product.objects.create(
                            sku=sku,
                            name=name,
                            description=description,
                            price=price,
                            cost=cost,
                            stock=stock_initial,
                            min_stock=int(min_stock) if min_stock is not None else 10,
                            max_stock=int(max_stock) if max_stock is not None else 1000,
                            category=category,
                            barcode=barcode,
                            tags=tags,
                            is_digital=is_digital,
                        )
                        
                        # Verificar que se creó correctamente
                        product.refresh_from_db()
                        logger.info(f"✅ Producto creado: SKU={sku}, ID={product.id}, Stock={product.stock} (verificado en BD)")
                        
                        if product.stock != stock_initial:
                            logger.error(f"⚠️ PROBLEMA: Stock no coincidente! Esperado: {stock_initial}, Obtenido: {product.stock}")
                    
                    successful += 1
                
                except Exception as e:
                    logger.error(f"❌ Error procesando fila {index + 2}: {str(e)}", exc_info=True)
                    errors.append(f"Fila {index + 2}: {str(e)}")
                    failed += 1
                    continue
        
        # La transacción se confirma automáticamente al salir del bloque with
        logger.info(f'💾 Transacción completada - Todos los cambios guardados en BD')
        
        logger.info(f'📊 ========== RESUMEN DE PROCESAMIENTO ==========')
        logger.info(f'📊 Total filas procesadas: {total_rows}')
        logger.info(f'✅ Exitosas: {successful}')
        logger.info(f'❌ Fallidas: {failed}')
        
        if errors:
            logger.warning(f'⚠️  Errores encontrados ({len(errors)}):')
            for error in errors[:10]:
                logger.warning(f'   - {error}')
        
        result = {
            'total_rows': total_rows,
            'successful': successful,
            'failed': failed,
            'errors': errors[:50]  # Limitar a 50 errores
        }
        
        logger.info(f'✅ ========== IMPORTACIÓN COMPLETADA ==========')
        
        result_serializer = ExcelImportResultSerializer(result)
        return Response(result_serializer.data, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f'❌ ========== ERROR EN IMPORTACIÓN ==========')
        logger.error(f'❌ Error: {str(e)}', exc_info=True)
        return Response({
            'error': f'Error al procesar archivo: {str(e)}',
            'traceback': traceback.format_exc() if request.user.is_staff else None
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def frequent_clients_report(request):
    """Reporte de clientes frecuentes"""
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        # Parámetros
        try:
            days = int(request.GET.get('days', 30))
            limit = int(request.GET.get('limit', 10))
            min_purchases = int(request.GET.get('min_purchases', 1))  # Cambiar a 1 mínimo
        except (ValueError, TypeError):
            days = 30
            limit = 10
            min_purchases = 1
        
        start_date = timezone.now() - timedelta(days=days)
        
        # Obtener todos los clientes activos que tengan ventas
        from apps.sales.models import Sale
        
        # Primero obtener IDs de clientes que tienen ventas
        client_ids_with_sales = Sale.objects.filter(
            status='completed',
            created_at__gte=start_date
        ).exclude(
            client__isnull=True
        ).values_list('client_id', flat=True).distinct()
        
        # Si no hay ventas en el período, devolver clientes ordenados por total_purchases del modelo
        if not client_ids_with_sales:
            logger.info('No hay ventas en el período especificado, devolviendo clientes por total de compras')
            # Anotar con datos de ventas históricas usando alias para evitar conflicto
            frequent_clients = Client.objects.filter(
                is_active=True
            ).annotate(
                purchase_count=Count('sale', filter=Q(sale__status='completed')),
                total_spent=Sum('sale__total', filter=Q(sale__status='completed')),
                last_sale_date=Max('sale__created_at', filter=Q(sale__status='completed'))  # Usar alias diferente
            ).order_by('-total_purchases', '-created_at')[:limit]
        else:
            # Obtener clientes con más compras en el período
            frequent_clients = Client.objects.filter(
                id__in=client_ids_with_sales,
                is_active=True
            ).annotate(
                purchase_count=Count('sale', filter=Q(sale__created_at__gte=start_date, sale__status='completed')),
                total_spent=Sum('sale__total', filter=Q(sale__created_at__gte=start_date, sale__status='completed')),
                last_sale_date=Max('sale__created_at', filter=Q(sale__status='completed'))  # Usar alias diferente
            ).filter(
                purchase_count__gte=min_purchases
            ).order_by('-purchase_count', '-total_spent')[:limit]
        
        serializer = FrequentClientSerializer(frequent_clients, many=True)
        
        # Estadísticas adicionales
        total_clients = Client.objects.filter(is_active=True).count()
        # Contar clientes que tienen al menos una venta completada
        clients_with_purchases = Client.objects.filter(
            is_active=True,
            sale__status='completed'
        ).distinct().count()
        
        return Response({
            'results': serializer.data,
            'period_days': days,
            'stats': {
                'total_clients': total_clients,
                'clients_with_purchases': clients_with_purchases,
                'frequent_clients_count': len(frequent_clients)
            }
        })
    except Exception as e:
        logger.error(f'Error en frequent_clients_report: {str(e)}', exc_info=True)
        return Response({
            'error': 'Error al cargar el reporte',
            'detail': str(e) if request.user.is_staff else 'Contacta al administrador'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_push_notification(request):
    """Enviar notificación push (admin only)"""
    import logging
    logger = logging.getLogger(__name__)
    
    if not request.user.is_staff:
        return Response({'error': 'No tienes permiso para enviar notificaciones'}, status=status.HTTP_403_FORBIDDEN)
    
    title = request.data.get('title')
    message = request.data.get('message')
    user_id = request.data.get('user_id')
    notification_type = request.data.get('notification_type', 'info')
    data = request.data.get('data', {})
    
    if not title or not message:
        return Response({'error': 'Título y mensaje son requeridos'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        from .services import ExpoPushNotificationService
        
        # Enviar a usuario específico o a todos
        if user_id:
            result = ExpoPushNotificationService.send_to_user_devices(
                user_id=user_id,
                title=title,
                message=message,
                data=data,
                notification_type=notification_type
            )
        else:
            # Enviar a todos los usuarios
            result = ExpoPushNotificationService.send_to_all_users(
                title=title,
                message=message,
                data=data,
                notification_type=notification_type
            )
        
        if result.get('success'):
            return Response({
                'message': f'Notificaciones enviadas: {result.get("sent", 0)}, Fallidas: {result.get("failed", 0)}',
                'sent': result.get('sent', 0),
                'failed': result.get('failed', 0),
                'total': result.get('total', 0)
            }, status=status.HTTP_200_OK)
        else:
            return Response({
                'error': result.get('message', 'Error al enviar notificaciones'),
                'sent': result.get('sent', 0),
                'failed': result.get('failed', 0)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
    except Exception as e:
        logger.error(f'Error enviando notificaciones push: {str(e)}', exc_info=True)
        return Response({
            'error': 'Error interno al enviar notificaciones',
            'details': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

