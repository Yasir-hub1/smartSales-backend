"""
Exportadores para diferentes formatos de reportes
"""
import os
import io
import csv
from datetime import datetime
from django.conf import settings
from django.http import HttpResponse
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch


class PDFExporter:
    """Exportador para archivos PDF"""
    
    @staticmethod
    def export_report(data, prompt):
        """Exporta un reporte a PDF"""
        try:
            # Crear nombre de archivo único
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'reporte_{timestamp}.pdf'
            
            # Crear buffer para el PDF
            buffer = io.BytesIO()
            
            # Crear documento PDF
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            styles = getSampleStyleSheet()
            
            # Contenido del PDF
            story = []
            
            # Título
            story.append(Paragraph("Reporte Generado", styles['Heading1']))
            story.append(Spacer(1, 12))
            
            # Información del reporte
            story.append(Paragraph(f"<b>Prompt:</b> {prompt}", styles['Normal']))
            story.append(Paragraph(f"<b>Generado:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Datos del reporte
            if isinstance(data, list) and len(data) > 0:
                # Crear tabla con los datos
                table_data = []
                
                # Encabezados
                if isinstance(data[0], dict):
                    headers = list(data[0].keys())
                    table_data.append(headers)
                    
                    # Datos
                    for row in data:
                        table_data.append([str(row.get(header, '')) for header in headers])
                else:
                    # Si no es una lista de diccionarios, mostrar como texto
                    story.append(Paragraph("<b>Datos:</b>", styles['Normal']))
                    story.append(Paragraph(str(data), styles['Normal']))
                
                if table_data:
                    # Crear tabla
                    table = Table(table_data)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 12),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    
                    story.append(table)
            else:
                story.append(Paragraph("No hay datos para mostrar", styles['Normal']))
            
            # Construir PDF
            doc.build(story)
            
            # Obtener contenido del buffer
            pdf_content = buffer.getvalue()
            buffer.close()
            
            # Guardar archivo
            file_path = f'reports/{filename}'
            default_storage.save(file_path, ContentFile(pdf_content))
            
            # Retornar URL del archivo
            return f'/media/{file_path}'
            
        except Exception as e:
            print(f"Error generando PDF: {str(e)}")
            return None


class ExcelExporter:
    """Exportador para archivos Excel"""
    
    @staticmethod
    def export_report(data, prompt):
        """Exporta un reporte a Excel"""
        try:
            # Crear nombre de archivo único
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'reporte_{timestamp}.xlsx'
            
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # Información del reporte
            ws['A1'] = "Reporte Generado"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:D1')
            
            ws['A3'] = f"Prompt: {prompt}"
            ws['A4'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            
            # Datos del reporte
            if isinstance(data, list) and len(data) > 0:
                # Encabezados
                if isinstance(data[0], dict):
                    headers = list(data[0].keys())
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=6, column=col, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                    
                    # Datos
                    for row_idx, row_data in enumerate(data, 7):
                        for col_idx, header in enumerate(headers, 1):
                            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))
                else:
                    # Si no es una lista de diccionarios, mostrar como texto
                    ws['A6'] = "Datos:"
                    ws['A7'] = str(data)
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar archivo
            file_path = f'reports/{filename}'
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            default_storage.save(file_path, ContentFile(buffer.getvalue()))
            buffer.close()
            
            # Retornar URL del archivo
            return f'/media/{file_path}'
            
        except Exception as e:
            print(f"Error generando Excel: {str(e)}")
            return None


class CSVExporter:
    """Exportador para archivos CSV"""
    
    @staticmethod
    def export_report(data, prompt):
        """Exporta un reporte a CSV"""
        try:
            # Crear nombre de archivo único
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'reporte_{timestamp}.csv'
            
            # Crear buffer para el CSV
            buffer = io.StringIO()
            
            if isinstance(data, list) and len(data) > 0:
                if isinstance(data[0], dict):
                    # Escribir CSV con diccionarios
                    fieldnames = list(data[0].keys())
                    writer = csv.DictWriter(buffer, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(data)
                else:
                    # Escribir CSV simple
                    writer = csv.writer(buffer)
                    for row in data:
                        writer.writerow([row])
            else:
                # Escribir datos simples
                buffer.write("No hay datos para mostrar")
            
            # Guardar archivo
            file_path = f'reports/{filename}'
            csv_content = buffer.getvalue()
            buffer.close()
            
            default_storage.save(file_path, ContentFile(csv_content.encode('utf-8')))
            
            # Retornar URL del archivo
            return f'/media/{file_path}'
            
        except Exception as e:
            print(f"Error generando CSV: {str(e)}")
            return None
